import openpyxl
import io
import logging

class ExcelService:
    @staticmethod
    def extract_text(excel_bytes: bytes) -> str:
        text_content = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
            for sheet in wb.worksheets:
                text_content.append(f"--- Sheet: {sheet.title} ---")
                for row in sheet.iter_rows(values_only=True):
                    # Filter out None values and convert to string
                    row_data = [str(cell) for cell in row if cell is not None]
                    if row_data:
                        text_content.append(" | ".join(row_data))
            from typing import Tuple, Optional
            return "\n".join(text_content), len([r for s in wb.worksheets for r in s.iter_rows(values_only=True) if any(cell is not None for cell in r)])
        except Exception as e:
            logging.error(f"Failed to read Excel file: {e}")
            return "", None
